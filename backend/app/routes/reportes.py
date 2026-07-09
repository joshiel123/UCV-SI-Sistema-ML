"""
Rutas para reportes y dashboard — Cable Latín
Generación optimizada de Excel con pandas/openpyxl en el servidor.
Protegidas con JWT y control de roles RBAC.
"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, g
from app.services.firebase_service import FirebaseService
from app.services.auth_middleware import token_required, role_required

# Intentar precargar librerías para evitar latencia de importación bajo demanda
try:
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    LIBRERIAS_DISPONIBLES = True
except ImportError:
    LIBRERIAS_DISPONIBLES = False

reportes_bp = Blueprint('reportes', __name__)
firebase_service = FirebaseService()

ROLES_SUPERVISION = ['supervisor', 'administrador']
TODOS_LOS_ROLES = ['vendedor', 'supervisor', 'administrador']


@reportes_bp.route('/dashboard', methods=['GET'])
@token_required
@role_required(TODOS_LOS_ROLES)
def dashboard():
    """Obtener datos de resumen para el tablero principal."""
    try:
        datos = firebase_service.obtener_resumen_dashboard()
        return jsonify({'success': True, 'data': datos}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': 'Error interno', 'errors': [str(e)]}), 500


@reportes_bp.route('/generar-excel', methods=['GET'])
@token_required
@role_required(ROLES_SUPERVISION)
def generar_excel():
    """
    Generar y descargar un reporte Excel completo con 3 pestañas:
    Clientes, Servicios/Facturas, Pagos de manera optimizada.

    Query params: ?fechaDesde=YYYY-MM-DD&fechaHasta=YYYY-MM-DD
    """
    if not LIBRERIAS_DISPONIBLES:
        return jsonify({
            'success': False,
            'message': 'Dependencias para reportes no instaladas en el servidor. Ejecute: pip install pandas openpyxl'
        }), 500

    try:
        fecha_desde = request.args.get('fechaDesde', '').strip()
        fecha_hasta = request.args.get('fechaHasta', '').strip()

        # Validar formato de fechas si se proveen
        for fecha_str, nombre in [(fecha_desde, 'fechaDesde'), (fecha_hasta, 'fechaHasta')]:
            if fecha_str:
                try:
                    datetime.strptime(fecha_str, '%Y-%m-%d')
                except ValueError:
                    return jsonify({
                        'success': False,
                        'message': f'Formato de fecha inválido para {nombre}. Use YYYY-MM-DD.'
                    }), 400

        # Obtener datos de Firestore de forma consolidada
        datos = firebase_service.obtener_datos_reporte(
            fecha_desde=fecha_desde or None,
            fecha_hasta=fecha_hasta or None
        )

        clientes = datos.get('clientes', [])
        servicios = datos.get('servicios', [])
        pagos = datos.get('pagos', [])

        # ===================== CONSTRUIR DATAFRAMES =====================

        # Estructuración eficiente de Clientes
        cols_clientes = ['nombre', 'apellido', 'tipo_documento', 'numero_documento',
                         'email', 'telefono', 'direccion', 'distrito', 'estado', 'fecha_registro']
        df_clientes = pd.DataFrame(clientes)
        df_clientes = df_clientes.reindex(columns=cols_clientes, fill_value='')
        df_clientes.columns = ['Nombre', 'Apellido', 'Tipo Doc.', 'Nro. Documento',
                               'Email', 'Teléfono', 'Dirección', 'Distrito', 'Estado', 'Fecha Registro']

        # Estructuración eficiente de Servicios/Facturas
        cols_servicios = ['id', 'cliente_id', 'plan_nombre', 'monto', 'estado',
                          'fecha_creacion', 'fecha_vencimiento', 'observacion']
        df_servicios = pd.DataFrame(servicios)
        df_servicios = df_servicios.reindex(columns=cols_servicios, fill_value='')
        df_servicios.columns = ['ID Servicio', 'ID Cliente', 'Plan', 'Monto (S/.)',
                                'Estado', 'Fecha Creación', 'Fecha Vencimiento', 'Observación']

        # Estructuración eficiente de Pagos
        cols_pagos = ['id', 'servicio_id', 'monto', 'metodo_pago', 'fecha_pago',
                      'procesado_por', 'observacion']
        df_pagos = pd.DataFrame(pagos)
        df_pagos = df_pagos.reindex(columns=cols_pagos, fill_value='')
        df_pagos.columns = ['ID Pago', 'ID Servicio', 'Monto (S/.)', 'Método de Pago',
                            'Fecha Pago', 'Procesado por', 'Observación']

        # ===================== GENERAR EXCEL EN MEMORIA =====================

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_clientes.to_excel(writer, sheet_name='Clientes', index=False, startrow=2)
            df_servicios.to_excel(writer, sheet_name='Servicios', index=False, startrow=2)
            df_pagos.to_excel(writer, sheet_name='Pagos', index=False, startrow=2)

            workbook = writer.book

            # Reutilización de estilos para mitigar sobrecarga de memoria
            header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            title_font = Font(bold=True, size=13, color='1A237E')
            alt_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
            border_side = Side(border_style='thin', color='BBBBBB')
            thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            
            # Instancia única de alineación de datos
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left_wrap = Alignment(vertical='center', wrap_text=True)

            rango_texto = _rango_fechas(fecha_desde, fecha_hasta)
            hojas_info = [
                ('Clientes', df_clientes, f'📋 Reporte de Clientes — Cable Latín | {rango_texto}'),
                ('Servicios', df_servicios, f'📄 Reporte de Servicios/Facturas — Cable Latín | {rango_texto}'),
                ('Pagos', df_pagos, f'💰 Reporte de Pagos — Cable Latín | {rango_texto}'),
            ]

            for nombre_hoja, df, titulo in hojas_info:
                ws = writer.sheets[nombre_hoja]

                # Título principal en la fila 1
                ws['A1'] = titulo
                ws['A1'].font = title_font
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[1].height = 22

                n_cols = len(df.columns)
                
                # Estilo de encabezados (Fila 3)
                for col_idx in range(1, n_cols + 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
                ws.row_dimensions[3].height = 20

                # Estilo de datos con filas alternadas (Fila 4 en adelante)
                for row_idx in range(4, len(df) + 4):
                    for col_idx in range(1, n_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill
                        cell.border = thin_border
                        cell.alignment = align_left_wrap

                # Auto-ajuste de columnas optimizado con Pandas (Operación Vectorizada)
                for col_idx, col_name in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    if not df.empty:
                        # Se calcula la longitud máxima usando operaciones nativas sobre la serie de Pandas
                        max_val_len = int(df[col_name].astype(str).str.len().max())
                        max_len = max(max_val_len, len(str(col_name)))
                    else:
                        max_len = len(str(col_name))
                        
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

                # Congelar panel superior (Encabezados estáticos al hacer scroll)
                ws.freeze_panes = 'A4'

        output.seek(0)

        # Nombre de salida único
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_archivo = f'Reporte_CableLatin_{fecha_actual}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno al compilar reporte: {str(e)}'}), 500


def _rango_fechas(fecha_desde: str, fecha_hasta: str) -> str:
    """Helper para estructurar el rango de fechas en los encabezados del reporte."""
    if fecha_desde and fecha_hasta:
        return f'{fecha_desde} al {fecha_hasta}'
    elif fecha_desde:
        return f'Desde {fecha_desde}'
    elif fecha_hasta:
        return f'Hasta {fecha_hasta}'
    return 'Historial Completo'